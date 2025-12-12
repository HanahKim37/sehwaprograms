import pandas as pd
from openpyxl import load_workbook

def load_seteuk(file):
    """세부능력 및 특기사항을 병합셀에서 추출하여 DF로 구성."""

    wb = load_workbook(file)
    ws = wb.active

    records = []

    # 1) 병합셀 순회
    for r in ws.merged_cells.ranges:
        (min_col, min_row, max_col, max_row) = r.bounds
        value = ws.cell(row=min_row, column=min_col).value

        # 세특 텍스트 후보 판별
        if not value:
            continue

        text = str(value).strip()
        if len(text) < 5:
            continue

        # 세특 내용 여부 (문장, 과목 구분 등)
        if ":" in text or "." in text:
            records.append({
                "row": min_row,
                "col": min_col,
                "세특내용": text
            })

    df = pd.DataFrame(records)

    if df.empty:
        return pd.DataFrame(columns=["번호", "성명", "학년", "세특내용"])

    # 2) 학생 정보 병합셀 추출 (번호, 성명, 학년)
    student_info = []

    for r in ws.merged_cells.ranges:
        (min_col, min_row, max_col, max_row) = r.bounds
        val = ws.cell(row=min_row, column=min_col).value
        if val is None:
            continue

        text = str(val)
        # 번호
        if text.isdigit():
            student_info.append({"row": min_row, "번호": text})
        # 성명 (한글 2~3자)
        elif len(text) in [2, 3] and all("가" <= ch <= "힣" for ch in text):
            student_info.append({"row": min_row, "성명": text})
        # 학년
        elif text in ["1", "2", "3"]:
            student_info.append({"row": min_row, "학년": text})

    df_info = pd.DataFrame(student_info)

    # 3) 학생정보와 세특내용 row 기반 매핑
    df["번호"] = None
    df["성명"] = None
    df["학년"] = None

    for i in df.index:
        row_pos = df.loc[i, "row"]
        info_subset = df_info[df_info["row"] <= row_pos]

        if not info_subset.empty:
            last_info = info_subset.sort_values("row").iloc[-1].to_dict()
            df.loc[i, "번호"] = last_info.get("번호", df.loc[i, "번호"])
            df.loc[i, "성명"] = last_info.get("성명", df.loc[i, "성명"])
            df.loc[i, "학년"] = last_info.get("학년", df.loc[i, "학년"])

    df["학년"] = pd.to_numeric(df["학년"], errors="coerce")

    return df[["번호", "성명", "학년", "세특내용"]].dropna(subset=["번호", "성명"])
